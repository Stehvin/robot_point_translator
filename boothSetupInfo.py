from openpyxl import load_workbook


def boothInfo(booth):
    """Find the parameters for a specific booth, which are listed in the
    excel document described by path (below).
    """

    # Input path to excel document.
    path = r'C:\Users\Example\UTool Records.xlsx'

    # Load excel workbook.
    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.get_sheet_by_name('Sheet1')
    except:
        return None

    # Initialize booth info dictionaries.
    tbl = {}
    utool = {}
    Euler = {}

    # Find proper row.
    for row in ws.iter_rows('A4:N{}'.format(ws.max_row)):
        if row[0].value == booth:
            
            # Get Utool set points.
            utool['x'] = row[2].value
            utool['y'] = row[3].value
            utool['z'] = row[4].value
            utool['w'] = 0
            utool['p'] = 0
            utool['r'] = 0
            
            # Get table center point.
            tbl['x'] = row[5].value
            tbl['y'] = row[6].value
            tbl['z'] = row[7].value

            # Get Euler angles.
            Euler['w'] = row[8].value
            Euler['p'] = row[9].value
            Euler['r'] = row[10].value

            # Copy dictionaries.
            utoolCopy = dict(utool)
            tblCopy = dict(tbl)
            EulerCopy = dict(Euler)

            # Get dust collector location and angle.
            dustColLoc = row[11].value
            dustColAng = row[12].value

            # Get configuration string.
            configStr = row[13].value
            break
    return utoolCopy, tblCopy, EulerCopy, dustColLoc, dustColAng, configStr


def curUtool(booth, curUtoolNum):
    """Returns the utool setpoints for the given booth and utool number.
    """
    # Input path to excel document.
    path = r'C:\Users\Example\UTool Records.xlsx'

    # Load excel workbook.
    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.get_sheet_by_name(str(booth))
    except:
        return None

    # Find current point/booth Utool.
    curUtool = {}
    for row in ws.iter_rows('A2:G{}'.format(ws.max_row)):
        if row[0].value == curUtoolNum:
            curUtool['x'] = row[1].value
            curUtool['y'] = row[2].value
            curUtool['z'] = row[3].value
            curUtool['w'] = row[4].value
            curUtool['p'] = row[5].value
            curUtool['r'] = row[6].value
            break
    return curUtool


def curUframe(booth, curUframeNum):
    """Returns the uframe setpoints for the given booth and uframe
    number.
    """
    # Input path to excel document.
    path = r'C:\Users\Example\UTool Records.xlsx'

    # Load excel workbook.
    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.get_sheet_by_name(str(booth))
    except:
        return None

    # Find current point/booth Utool.
    curUframe = {}
    for row in ws.iter_rows('K2:Q{}'.format(ws.max_row)):
        if row[0].value == curUframeNum:
            curUframe['x'] = row[1].value
            curUframe['y'] = row[2].value
            curUframe['z'] = row[3].value
            curUframe['w'] = row[4].value
            curUframe['p'] = row[5].value
            curUframe['r'] = row[6].value
            break
    return curUframe
